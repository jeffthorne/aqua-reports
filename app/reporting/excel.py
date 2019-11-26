from typing import List, Set, Dict, Tuple, Optional
import time
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from app.models.models import Image


class ExcelReport:

    def __init__(self, image: Image, path: Path):
        self.image = image
        self.path = path


    def generate(self):
        workbook = Workbook()
        self._risk_sheet(workbook)
        self.__vuln_sheet(workbook)
        self.__sensitive_data_sheet(workbook)
        self.__malware_datasheet(workbook)

        workbook.remove_sheet(workbook.get_sheet_by_name('Sheet'))
        workbook.save(filename=f"{str(self.path.absolute())}/{self.image.registry.replace(' ', '-')}_{self.image.repository.replace('/', '_')}_{self.image.tag}_{time.strftime('%Y%m%d-%H%M%S')}.xlsx")

    def __setup_worksheet_columns(self, worksheet, columns):
        for index, column in enumerate(columns):
            worksheet.cell(1, index + 1).value = column

    def __malware_datasheet(self, workbook):
        mal_sheet = workbook.create_sheet(title="Malware")
        columns = ['Malware', 'Hash', 'Path', 'Paths', 'Aknowledged', 'Aknowledge Date', 'Aknowledge Scope']
        self.__setup_worksheet_columns(mal_sheet, columns)
        fields_to_parse = [x.lower().replace(" ", "_")  for x in columns]

        if self.image.malware_findings is not None:
            self.__add_cell_to_worksheet(mal_sheet, 2, fields_to_parse, self.image.malware_findings)

        self.__bold_row(mal_sheet)
        self.__col_width(mal_sheet, 30)
        self.__left_cells(mal_sheet)


    def __add_cell_to_worksheet(self, worksheet, starting_row: int, cell_names: List, json_object: Dict):

        for row in json_object:
            for i, name in enumerate(cell_names):
                worksheet.cell(starting_row, i + 1).value = self.__parse_celldata_for_worksheet(name, row)

            starting_row = starting_row + 1



    def __sensitive_data_sheet(self, workbook):
        sd_sheet = workbook.create_sheet(title="Sensitive Data")
        columns = ['Type', 'Path', 'Hash', 'Filename', 'Aknowledged', 'Aknowledge Date', 'Aknowledge Scope']
        fields_to_parse = [x.lower().replace(" ", "_") for x in columns]
        self.__setup_worksheet_columns(sd_sheet, columns)

        if self.image.sensitive_data_results is not None:
            self.__add_cell_to_worksheet(sd_sheet, 2, fields_to_parse, self.image.sensitive_data_results)

        self.__bold_row(sd_sheet)
        self.__col_width(sd_sheet, 30)
        self.__left_cells(sd_sheet)



    def __vuln_sheet(self, workbook):
        vuln_sheet = workbook.create_sheet(title="Vulnerabilities")
        columns = ['Vulnerability', 'Resource', 'Installed Version', 'Fix Version', 'Solution', 'Aqua Score', 'Aqua Severity', 'NVD Severity',
                   'NVD Score', 'NVD Vectors', 'NVD CVSS v3 Severity', 'NVD CVSS v3 Score', 'NVD CVSS v3 Vectors', 'NVD Reference',
                   'Vendor Score', 'Vendor Severity', 'Vendor Vectors', 'Vendor Reference', 'Publish Date',
                   'Modification Date', 'Description']

        self.__setup_worksheet_columns(vuln_sheet, columns)

        if self.image.vulnerabilities is not None:
            row = 2
            for vuln in self.image.vulnerabilities:
                vuln_sheet.cell(row, 1).value = vuln['name']
                vuln_sheet.cell(row, 2).value = str(vuln['resource'])
                vuln_sheet.cell(row, 3).value = vuln['resource']['version'] if vuln['resource']['version'] != "" else "-"
                vuln_sheet.cell(row, 4).value = vuln['fix_version'] if vuln['fix_version'] != "" else "None"
                vuln_sheet.cell(row, 5).value = vuln['solution']
                vuln_sheet.cell(row, 6).value = vuln['aqua_score']
                vuln_sheet.cell(row, 7).value = vuln['aqua_severity']
                vuln_sheet.cell(row, 8).value = vuln['nvd_severity']
                vuln_sheet.cell(row, 8).value = vuln['nvd_cvss2_score']
                vuln_sheet.cell(row, 10).value = vuln['nvd_cvss2_vectors']

                vuln_sheet.cell(row, 11).value = vuln['nvd_cvss3_severity'] if vuln['nvd_cvss3_severity'] != "" else "-"
                vuln_sheet.cell(row, 12).value = vuln['nvd_cvss3_score'] if vuln['nvd_cvss3_score'] != "" else "-"
                vuln_sheet.cell(row, 13).value = vuln['nvd_cvss3_vectors'] if vuln['nvd_cvss3_vectors'] != "" else "-"

                vuln_sheet.cell(row, 14).value = vuln['nvd_url'] if vuln['nvd_url'] != "" else "-"
                vuln_sheet.cell(row, 15).value = self.__parse_celldata_for_worksheet('vendor_cvss3_score', vuln)
                vuln_sheet.cell(row, 16).value = vuln['vendor_severity'] if vuln['vendor_severity'] != "" else "-"
                vuln_sheet.cell(row, 17).value = self.__parse_celldata_for_worksheet('vendor_cvss3_vectors', vuln)
                vuln_sheet.cell(row, 18).value = vuln['vendor_url'] if vuln['vendor_url'] != "" else "-"
                vuln_sheet.cell(row, 19).value = vuln['publish_date'] if vuln['publish_date'] != "" else "-"
                vuln_sheet.cell(row, 20).value = vuln['modification_date'] if vuln['modification_date'] != "" else "-"
                vuln_sheet.cell(row, 21).value = vuln['description']
                row = row + 1

        self.__bold_row(vuln_sheet)
        self.__col_width(vuln_sheet, 20)
        self.__left_cells(vuln_sheet)

    def __parse_celldata_for_worksheet(self, name, json_obj):
        if name in json_obj and json_obj[name] != "":
            if type(json_obj[name]) is list:
                return ', '.join(json_obj[name])
            else:
                return json_obj[name]
        else:
            return "-"


    def _risk_sheet(self, workbook):
        risk_sheet = workbook.create_sheet(title="Risk")
        risk_columns = ['Image', 'Tag', 'Registry', 'Digest', 'Compliant', 'Scan Date', 'Total Vulns', 'Critical', 'High', 'Medium', 'Low',
                        'Negligible', 'Malware Count', 'Sensitive Data Count', 'Whitelisted', 'Blacklisted']


        self.__setup_worksheet_columns(risk_sheet, risk_columns)

        row = 2
        risk_sheet.cell(row, 1).value = self.image.name
        risk_sheet.cell(row, 2).value = self.image.tag
        risk_sheet.cell(row, 3).value = self.image.registry
        risk_sheet.cell(row, 4).value = self.image.digest
        risk_sheet.cell(row, 5).value = self.image.disallowed
        risk_sheet.cell(row, 6).value = self.image.scan_date
        risk_sheet.cell(row, 7).value = self.image.vulns_found
        risk_sheet.cell(row, 8).value = self.image.crit_vulns
        risk_sheet.cell(row, 9).value = self.image.high_vulns
        risk_sheet.cell(row, 10).value = self.image.med_vulns
        risk_sheet.cell(row, 11).value = self.image.low_vulns
        risk_sheet.cell(row, 12).value = self.image.neg_vulns
        risk_sheet.cell(row, 13).value = self.image.malware
        risk_sheet.cell(row, 14).value = self.image.sensitive_data
        risk_sheet.cell(row, 15).value = self.image.whitelisted
        risk_sheet.cell(row, 16).value = self.image.blacklisted

        self.__bold_row(risk_sheet)
        self.__col_width(risk_sheet, 20)
        self.__center_cells(risk_sheet)

    def __bold_row(self, ws):
        for cell in ws[1]:
            cell.font = Font(bold=True)

    def __col_width(self, ws, max):
        max_length = 0
        for col in ws.columns:
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width if adjusted_width <= max else max


    def __center_cells(self, ws):
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)


    def __left_cells(self, ws):
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

